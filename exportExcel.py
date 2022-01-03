"""
This file is in charge of exporting the data and convert them into an excel file.
External library used: - openpyxl -
"""

from openpyxl import Workbook
from database import getAllUsers, getScore, getMessage


def exportScores(userID):
    """ Convert, calculate and export the data into an excel file. """
    data = getScore(userID)
    if data is None:
        return

    sheet = Workbook()  # The excel worksheet

    gameCatch = {
        'count': 0,  # Amount of times the player played that game
        'min': 0,  # Minimum score
        'max': 0,  # Maximum score
        'sum': 0  # Sum of score
    }
    gameSize = {
        'count': 0,  # Amount of times the player played that game
        'min': 0,  # Minimum score
        'max': 0,  # Maximum score
        'sum': 0  # Sum of score
    }
    gameCount = {
        'count': 0,  # Amount of times the player played that game
        'min': 0,  # Minimum score
        'max': 0,  # Maximum score
        'sum': 0  # Sum of score
    }
    gameMathExp = {
        'count': 0,  # Amount of times the player played that game
        'min': 0,  # Minimum score
        'max': 0,  # Maximum score
        'sum': 0  # Sum of score
    }

    # Map the data into the variables
    for d in data:
        score = d['score']
        if d['type'] == "Choose Size":
            gameSize['count'] += 1
            gameSize['sum'] += score

            if gameSize['count'] == 1:
                gameSize['min'] = score
                gameSize['max'] = score
            else:
                gameSize['min'] = min(gameSize['min'], score)
                gameSize['max'] = max(gameSize['min'], score)
        elif d['type'] == "Count Game":
            gameCount['count'] += 1
            gameCount['sum'] += score

            if gameCount['count'] == 1:
                gameCount['min'] = score
                gameCount['max'] = score
            else:
                gameCount['min'] = min(gameCount['min'], score)
                gameCount['max'] = max(gameCount['min'], score)
        elif d['type'] == "Catch the Answer":
            gameCatch['count'] += 1
            gameCatch['sum'] += score

            if gameCatch['count'] == 1:
                gameCatch['min'] = score
                gameCatch['max'] = score
            else:
                gameCatch['min'] = min(gameCatch['min'], score)
                gameCatch['max'] = max(gameCatch['min'], score)
        else:
            gameMathExp['count'] += 1
            gameMathExp['sum'] += score

            if gameMathExp['count'] == 1:
                gameMathExp['min'] = score
                gameMathExp['max'] = score
            else:
                gameMathExp['min'] = min(gameMathExp['min'], score)
                gameMathExp['max'] = max(gameMathExp['min'], score)

    gameSize['avg'] = gameSize['sum'] / max(gameSize['count'], 1)
    gameCount['avg'] = gameCount['sum'] / max(gameCount['count'], 1)
    gameCatch['avg'] = gameCatch['sum'] / max(gameCatch['count'], 1)
    gameMathExp['avg'] = gameMathExp['sum'] / max(gameMathExp['count'], 1)

    # Convert the data into table format
    content = (
        ("Game Type:", "Catch the Answer", "Math Expressions", "SizeMe", "Count Game"),
        ("Amount of plays", gameCatch['count'], gameMathExp['count'], gameSize['count'], gameCount['count']),
        ("Best Score", gameCatch['max'], gameMathExp['max'], gameSize['max'], gameCount['max']),
        ("Worst Score", gameCatch['min'], gameMathExp['min'], gameSize['min'], gameCount['min']),
        ("Average Score", gameCatch['avg'], gameMathExp['avg'], gameSize['avg'], gameCount['avg'])
    )

    # Adding the data into the excel object
    for c in content:
        sheet.active.append(c)
    sheet.save('{} score report.xlsx'.format(userID))

    sheet.close()


def exportMessages(userID):
    """ Convert, calculate and export the data into an excel file. """
    data = getMessage(userID)
    if data is None:
        return

    sheet = Workbook()

    # Adding the data into the excel object
    sheetWrite = sheet.active
    sheetWrite.append(("Date", "From", "Message"))
    for message in data:
        sheetWrite.append((message[1]['date'], message[1]['from'], message[1]['info']))
    sheet.save('Messages for {}.xlsx'.format(userID))

    sheet.close()


def exportUsers():
    """ Convert, calculate and export the data into an excel file. """
    users = getAllUsers()  # Get all users from the database
    if users is None:
        return

    sheet = Workbook()  # The excel worksheet
    children = []
    parents = []
    diagnostics = []

    # Map the data into the corresponding lists
    for user in users.items():
        if user[1]['type'] == "Child":
            children.append((user[0], user[1]['last'], user[1]['name']))
        elif user[1]['type'] == "Parent":
            parents.append((user[0], user[1]['last'], user[1]['name']))
        else:
            diagnostics.append((user[0], user[1]['last'], user[1]['name']))

    # Adding the data into the excel object
    sheetWrite = sheet.active
    sheetWrite.append(("Account Type", "ID", "Last Name", "First Name"))
    sheetWrite.append(("Children", ""))
    for content in children:
        sheetWrite.append(("", content[0], content[1], content[2]))
    sheetWrite.append(("Parents", ""))
    for content in parents:
        sheetWrite.append(("", content[0], content[1], content[2]))
    sheetWrite.append(("Diagnostics", ""))
    for content in diagnostics:
        sheetWrite.append(("", content[0], content[1], content[2]))
    sheet.save('all users report.xlsx')

    sheet.close()
